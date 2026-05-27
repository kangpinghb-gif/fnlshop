from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
import csv

from openpyxl import load_workbook


ALIASES = {
    "store_code": ["store_code", "门店编码", "店铺编号", "门店编号"],
    "store_name": ["store_name", "门店名称", "店铺名称"],
    "product_code": ["product_code", "商品编码", "产品编码"],
    "product_name": ["product_name", "商品名称", "产品名称", "品名"],
    "count_time": ["count_time", "盘点时间", "修正时间"],
    "adjusted_quantity": ["adjusted_quantity", "人工盘点修正值", "盘点修正值", "盘点数量"],
    "unit": ["unit", "单位", "计量单位"],
    "count_type": ["count_type", "盘点类型"],
    "remark": ["remark", "备注", "说明"],
}


@dataclass(frozen=True)
class StockAdjustmentRow:
    source_file_name: str
    source_sheet_name: str
    raw_row_number: int
    store_code: str
    store_name: str
    product_code: str
    product_name: str
    count_time: str
    business_date: str
    adjusted_quantity: float
    unit: str
    count_type: str
    remark: str

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


def parse_stock_adjustment_file(
    path: str | Path,
    *,
    default_business_date: str = "",
) -> list[StockAdjustmentRow]:
    input_path = Path(path)
    if input_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(input_path, default_business_date=default_business_date)
    return _read_csv_rows(input_path, default_business_date=default_business_date)


def _read_csv_rows(path: Path, *, default_business_date: str) -> list[StockAdjustmentRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            return []
        field_map = _build_field_map(reader.fieldnames)
        rows = []
        for raw_row_number, raw in enumerate(reader, start=2):
            row = _parse_row(
                raw,
                field_map,
                source_file_name=path.name,
                source_sheet_name="",
                raw_row_number=raw_row_number,
                default_business_date=default_business_date,
            )
            if row:
                rows.append(row)
        return rows


def _read_xlsx_rows(path: Path, *, default_business_date: str) -> list[StockAdjustmentRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[StockAdjustmentRow] = []
    try:
        for sheet in workbook.worksheets:
            values_iter = sheet.iter_rows(values_only=True)
            headers = next(values_iter, None)
            if not headers:
                continue
            fieldnames = [_clean_text(value) for value in headers]
            field_map = _build_field_map(fieldnames)
            if "adjusted_quantity" not in field_map:
                continue
            for raw_row_number, values in enumerate(values_iter, start=2):
                raw = dict(zip(fieldnames, values))
                row = _parse_row(
                    raw,
                    field_map,
                    source_file_name=path.name,
                    source_sheet_name=sheet.title,
                    raw_row_number=raw_row_number,
                    default_business_date=default_business_date,
                )
                if row:
                    rows.append(row)
    finally:
        workbook.close()
    return rows


def _parse_row(
    raw: dict[str, object],
    field_map: dict[str, str],
    *,
    source_file_name: str,
    source_sheet_name: str,
    raw_row_number: int,
    default_business_date: str,
) -> StockAdjustmentRow | None:
    adjusted_quantity = _to_float(_get(raw, field_map, "adjusted_quantity"))
    if adjusted_quantity is None:
        return None

    count_time = _to_datetime_text(_get(raw, field_map, "count_time"))
    business_date = _to_date_text(count_time[:10]) or _to_date_text(default_business_date)
    if not count_time and business_date:
        count_time = f"{business_date} 00:00:00"
    if not business_date:
        return None

    return StockAdjustmentRow(
        source_file_name=source_file_name,
        source_sheet_name=source_sheet_name,
        raw_row_number=raw_row_number,
        store_code=_clean_code(_get(raw, field_map, "store_code")),
        store_name=_get(raw, field_map, "store_name"),
        product_code=_clean_code(_get(raw, field_map, "product_code")),
        product_name=_get(raw, field_map, "product_name"),
        count_time=count_time,
        business_date=business_date,
        adjusted_quantity=adjusted_quantity,
        unit=_get(raw, field_map, "unit") or "kg",
        count_type=_get(raw, field_map, "count_type") or "manual",
        remark=_get(raw, field_map, "remark"),
    )


def _build_field_map(fieldnames: list[str]) -> dict[str, str]:
    normalized = {_normalize_header(name): name for name in fieldnames}
    result: dict[str, str] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            source = normalized.get(_normalize_header(alias))
            if source:
                result[field] = source
                break
    return result


def _get(raw: dict[str, object], field_map: dict[str, str], field: str) -> str:
    source = field_map.get(field)
    if not source:
        return ""
    return _clean_text(raw.get(source))


def _normalize_header(value: object) -> str:
    return _clean_text(value).replace(" ", "").replace("\n", "").replace("\t", "")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "-"} else text


def _clean_code(value: object) -> str:
    text = _clean_text(value)
    return text[:-2] if text.endswith(".0") else text


def _to_float(value: object) -> float | None:
    text = _clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_date_text(value: object) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d", "%m-%d", "%m.%d"):
        try:
            parsed = datetime.strptime(text, fmt)
        except ValueError:
            continue
        year = parsed.year if "%Y" in fmt else date.today().year
        return date(year, parsed.month, parsed.day).isoformat()
    return text


def _to_datetime_text(value: object) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text

from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
import csv
from collections.abc import Iterable

from openpyxl import load_workbook


REPORT_TYPES = {"sales", "inventory_loss", "purchase_receipts", "inventory_snapshot"}


ALIASES = {
    "store_code": ["店铺编号", "门店编号", "店铺编码", "门店编码"],
    "product_code": ["商品编码", "产品编码"],
    "business_date": ["日期", "营业日期", "业务日期", "销售日期", "发生日期"],
    "sales_quantity": ["销售数量", "销量", "销售量"],
    "sales_amount": ["销售金额", "销售额", "销售收入"],
    "unit": ["单位", "销售单位", "计量单位"],
    "closing_stock_qty": ["库存数量", "期末库存", "期末库存数量", "库存数量（期末）"],
    "loss_quantity": ["报损数量", "损耗数量"],
    "loss_amount": ["报损金额", "损耗金额"],
    "inventory_difference_qty": ["盘盈盘亏数量", "盘点差异数量", "库存差异数量"],
    "order_quantity": ["订货数量", "订单数量"],
    "receive_quantity": ["收货数量", "入库数量"],
    "total_receive_quantity": ["总收货数量", "总入库数量"],
    "total_return_quantity": ["总退货+调出数量", "退货数量", "调出数量"],
    "inventory_quantity": ["实时库存", "实时库存数量", "库存数量", "门店库存数量"],
    "snapshot_time": ["快照时间", "导出时间", "库存时间"],
}


@dataclass(frozen=True)
class DabiaogeDailyRow:
    report_type: str
    store_code: str
    product_code: str
    business_date: str
    unit: str
    sales_quantity: float | None = None
    sales_amount: float | None = None
    closing_stock_qty: float | None = None
    loss_quantity: float | None = None
    loss_amount: float | None = None
    inventory_difference_qty: float | None = None
    order_quantity: float | None = None
    receive_quantity: float | None = None
    total_receive_quantity: float | None = None
    total_return_quantity: float | None = None
    inventory_quantity: float | None = None
    snapshot_time: str = ""
    inventory_source: str = ""
    source_file: str = ""

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


def parse_dabiaoge_daily_csv(
    path: str | Path,
    *,
    report_type: str,
    default_business_date: str = "",
) -> list[DabiaogeDailyRow]:
    return parse_dabiaoge_daily_file(
        path,
        report_type=report_type,
        default_business_date=default_business_date,
    )


def parse_dabiaoge_daily_file(
    path: str | Path,
    *,
    report_type: str,
    default_business_date: str = "",
) -> list[DabiaogeDailyRow]:
    if report_type not in REPORT_TYPES:
        raise ValueError(f"Unsupported report_type: {report_type}")

    input_path = Path(path)
    if input_path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(input_path, report_type=report_type, default_business_date=default_business_date)
    return _read_csv_rows(input_path, report_type=report_type, default_business_date=default_business_date)


def merge_dabiaoge_daily_rows(rows: Iterable[DabiaogeDailyRow]) -> list[DabiaogeDailyRow]:
    merged: dict[tuple[str, str, str, str], DabiaogeDailyRow] = {}
    for row in rows:
        key = (row.report_type, row.store_code, row.product_code, row.business_date)
        existing = merged.get(key)
        if not existing:
            merged[key] = row
            continue
        merged[key] = DabiaogeDailyRow(
            report_type=row.report_type,
            store_code=row.store_code,
            product_code=row.product_code,
            business_date=row.business_date,
            unit=row.unit or existing.unit,
            sales_quantity=_sum_optional(existing.sales_quantity, row.sales_quantity),
            sales_amount=_sum_optional(existing.sales_amount, row.sales_amount),
            closing_stock_qty=_latest_optional(existing.closing_stock_qty, row.closing_stock_qty),
            loss_quantity=_sum_optional(existing.loss_quantity, row.loss_quantity),
            loss_amount=_sum_optional(existing.loss_amount, row.loss_amount),
            inventory_difference_qty=_sum_optional(existing.inventory_difference_qty, row.inventory_difference_qty),
            order_quantity=_sum_optional(existing.order_quantity, row.order_quantity),
            receive_quantity=_sum_optional(existing.receive_quantity, row.receive_quantity),
            total_receive_quantity=_sum_optional(existing.total_receive_quantity, row.total_receive_quantity),
            total_return_quantity=_sum_optional(existing.total_return_quantity, row.total_return_quantity),
            inventory_quantity=_latest_optional(existing.inventory_quantity, row.inventory_quantity),
            snapshot_time=row.snapshot_time or existing.snapshot_time,
            inventory_source=row.inventory_source or existing.inventory_source,
            source_file=_merge_source_files(existing.source_file, row.source_file),
        )
    return list(merged.values())


def _read_csv_rows(
    path: Path,
    *,
    report_type: str,
    default_business_date: str,
) -> list[DabiaogeDailyRow]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if not reader.fieldnames:
            return []
        field_map = _build_field_map(reader.fieldnames)
        rows = []
        for raw in reader:
            row = _parse_row(
                raw,
                field_map,
                report_type=report_type,
                source_file=path.name,
                default_business_date=default_business_date,
            )
            if row:
                rows.append(row)
        return rows


def _read_xlsx_rows(
    path: Path,
    *,
    report_type: str,
    default_business_date: str,
) -> list[DabiaogeDailyRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[DabiaogeDailyRow] = []
    try:
        for sheet in workbook.worksheets:
            values_iter = sheet.iter_rows(values_only=True)
            headers = next(values_iter, None)
            if not headers:
                continue
            fieldnames = [_clean_text(value) for value in headers]
            field_map = _build_field_map(fieldnames)
            if not field_map:
                continue
            for values in values_iter:
                raw = dict(zip(fieldnames, values))
                row = _parse_row(
                    raw,
                    field_map,
                    report_type=report_type,
                    source_file=f"{path.name}:{sheet.title}",
                    default_business_date=default_business_date,
                )
                if row:
                    rows.append(row)
    finally:
        workbook.close()
    return rows


def _parse_row(
    raw: dict[str, object],
    field_map: dict[str, str],
    *,
    report_type: str,
    source_file: str,
    default_business_date: str,
) -> DabiaogeDailyRow | None:
    store_code = _clean_code(_get(raw, field_map, "store_code"))
    product_code = _clean_code(_get(raw, field_map, "product_code"))
    business_date = _to_date_text(_get(raw, field_map, "business_date")) or _to_date_text(default_business_date)
    if not store_code or not product_code or not business_date:
        return None

    return DabiaogeDailyRow(
        report_type=report_type,
        store_code=store_code,
        product_code=product_code,
        business_date=business_date,
        unit=_get(raw, field_map, "unit") or "kg",
        sales_quantity=_to_float(_get(raw, field_map, "sales_quantity")),
        sales_amount=_to_float(_get(raw, field_map, "sales_amount")),
        closing_stock_qty=_to_float(_get(raw, field_map, "closing_stock_qty")),
        loss_quantity=_to_float(_get(raw, field_map, "loss_quantity")),
        loss_amount=_to_float(_get(raw, field_map, "loss_amount")),
        inventory_difference_qty=_to_float(_get(raw, field_map, "inventory_difference_qty")),
        order_quantity=_to_float(_get(raw, field_map, "order_quantity")),
        receive_quantity=_to_float(_get(raw, field_map, "receive_quantity")),
        total_receive_quantity=_to_float(_get(raw, field_map, "total_receive_quantity")),
        total_return_quantity=_to_float(_get(raw, field_map, "total_return_quantity")),
        inventory_quantity=_to_float(_get(raw, field_map, "inventory_quantity")),
        snapshot_time=_to_datetime_text(_get(raw, field_map, "snapshot_time")) or f"{business_date} 00:00:00",
        inventory_source="realtime" if report_type == "inventory_snapshot" else "",
        source_file=source_file,
    )


def _build_field_map(fieldnames: list[str]) -> dict[str, str]:
    normalized = {_normalize_header(name): name for name in fieldnames}
    result: dict[str, str] = {}
    for field, aliases in ALIASES.items():
        for alias in aliases:
            source = normalized.get(_normalize_header(alias))
            if source:
                result[field] = source
                break
    return result


def _get(raw: dict[str, object], field_map: dict[str, str], field: str) -> str:
    source = field_map.get(field)
    if not source:
        return ""
    return _clean_text(raw.get(source))


def _normalize_header(value: object) -> str:
    return _clean_text(value).replace(" ", "").replace("\n", "").replace("\t", "")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "-"} else text


def _clean_code(value: object) -> str:
    text = _clean_text(value)
    return text[:-2] if text.endswith(".0") else text


def _to_float(value: object) -> float | None:
    text = _clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _sum_optional(left: float | None, right: float | None) -> float | None:
    if left is None:
        return right
    if right is None:
        return left
    return left + right


def _latest_optional(left: float | None, right: float | None) -> float | None:
    return right if right is not None else left


def _merge_source_files(left: str, right: str) -> str:
    if not left:
        return right
    if not right or right == left:
        return left
    sources = left.split("+")
    if right in sources:
        return left
    return f"{left}+{right}"


def _to_date_text(value: object) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%m/%d", "%m-%d", "%m.%d"):
        try:
            parsed = datetime.strptime(text, fmt)
        except ValueError:
            continue
        year = parsed.year if "%Y" in fmt else date.today().year
        return date(year, parsed.month, parsed.day).isoformat()
    return text


def _to_datetime_text(value: object) -> str:
    text = _clean_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            continue
    return text

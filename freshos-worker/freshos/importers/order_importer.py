from __future__ import annotations

from dataclasses import asdict, dataclass
from datetime import date, datetime
from pathlib import Path
import re
from typing import Iterable

from openpyxl import load_workbook


ARRIVAL_QTY_PRIORITY = ["门店实收重量", "净果数量", "发货数量", "供货量", "订货数量"]

HEADER_ALIASES = {
    "supplier_code": ["供应商编码", "供商编码"],
    "supplier_name": ["供应商", "供应商名称", "供货商", "供货商名称"],
    "store_name_raw": ["门店", "门店名称", "店铺", "店铺名称", "收货门店"],
    "product_name_raw": ["商品", "商品名称", "品名", "产品名称", "订单商品名称"],
    "order_date": ["订单日期", "订货日期", "下单日期"],
    "arrival_date": ["到货日期", "发货日期", "配送日期", "送货日期"],
    "ordered_quantity": ["订货数量", "订单数量", "下单数量", "要货数量", "订单量", "订单量（kg）"],
    "arrival_quantity": ["到货数量", "入库数量", "实收数量"],
    "gross_quantity": ["毛重", "总重", "毛重数量", "毛重数量（kg）"],
    "tare_quantity": ["皮重", "筐皮", "筐重", "筐皮数量", "筐皮数量（kg）"],
    "received_quantity": ["门店实收重量", "实收重量", "门店实收重量（kg）"],
    "net_quantity": ["净果数量", "净重", "净果数量（kg）"],
    "shipped_quantity": ["发货数量", "发出数量"],
    "supplied_quantity": ["供货量", "供货数量"],
    "unit": ["单位", "销售单位", "计量单位"],
    "remark": ["备注", "说明"],
}


@dataclass(frozen=True)
class FreshOrderImportRow:
    source_file_name: str
    source_sheet_name: str
    source_format: str
    supplier_code: str
    supplier_name: str
    store_name_raw: str
    product_name_raw: str
    order_date: str
    arrival_date: str
    ordered_quantity: float | None
    arrival_quantity: float
    gross_quantity: float | None
    tare_quantity: float | None
    received_quantity: float | None
    unit: str
    match_status: str
    remark: str
    raw_row_number: int

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


def parse_order_workbook(
    path: str | Path,
    *,
    source_format: str = "auto",
    arrival_date_override: date | None = None,
    order_date_override: date | None = None,
) -> list[FreshOrderImportRow]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=True)
    rows: list[FreshOrderImportRow] = []
    for sheet in wb.worksheets:
        parsed = _parse_sheet(
            workbook_path,
            sheet,
            source_format=source_format,
            arrival_date_override=arrival_date_override,
            order_date_override=order_date_override,
        )
        rows.extend(parsed)
    return rows


def rows_to_dicts(rows: Iterable[FreshOrderImportRow]) -> list[dict[str, object]]:
    return [row.to_dict() for row in rows]


def _parse_sheet(
    workbook_path: Path,
    sheet,
    *,
    source_format: str,
    arrival_date_override: date | None,
    order_date_override: date | None,
) -> list[FreshOrderImportRow]:
    header_row_idx, field_to_col = _detect_header(sheet)
    if header_row_idx is None:
        return []

    metadata = _sheet_metadata(workbook_path, sheet)
    default_arrival_date = arrival_date_override or metadata.get("arrival_date") or _date_from_filename(workbook_path.name)
    default_order_date = order_date_override or default_arrival_date
    fallback_store_name = metadata.get("store_name") or _store_name_from_filename_or_sheet(workbook_path, sheet.title)
    default_supplier_code = metadata.get("supplier_code") or ""
    default_supplier_name = metadata.get("supplier_name") or ""

    parsed: list[FreshOrderImportRow] = []
    blank_streak = 0
    for row_idx in range(header_row_idx + 1, sheet.max_row + 1):
        values = {field: _cell_value(sheet.cell(row_idx, col_idx).value) for field, col_idx in field_to_col.items()}
        product_name = values.get("product_name_raw", "")
        if not product_name:
            blank_streak += 1
            if blank_streak >= 50:
                break
            continue
        blank_streak = 0

        store_name = values.get("store_name_raw") or fallback_store_name
        if not store_name:
            continue

        order_dt = _coerce_date(values.get("order_date")) or default_order_date
        arrival_dt = _coerce_date(values.get("arrival_date")) or default_arrival_date
        if arrival_dt is None:
            continue

        ordered_quantity = _to_float(values.get("ordered_quantity"))
        arrival_quantity = _choose_arrival_quantity(values, ordered_quantity)
        if arrival_quantity is None:
            continue

        parsed.append(
            FreshOrderImportRow(
                source_file_name=workbook_path.name,
                source_sheet_name=sheet.title,
                source_format=_resolve_source_format(source_format, field_to_col, workbook_path, metadata),
                supplier_code=values.get("supplier_code", "") or default_supplier_code,
                supplier_name=values.get("supplier_name", "") or default_supplier_name,
                store_name_raw=store_name,
                product_name_raw=product_name,
                order_date=order_dt.isoformat() if order_dt else "",
                arrival_date=arrival_dt.isoformat(),
                ordered_quantity=ordered_quantity,
                arrival_quantity=arrival_quantity,
                gross_quantity=_to_float(values.get("gross_quantity")),
                tare_quantity=_to_float(values.get("tare_quantity")),
                received_quantity=_to_float(values.get("received_quantity")),
                unit=values.get("unit") or "kg",
                match_status="pending",
                remark=values.get("remark", ""),
                raw_row_number=row_idx,
            )
        )
    return parsed


def _detect_header(sheet) -> tuple[int | None, dict[str, int]]:
    for row_idx in range(1, min(sheet.max_row, 30) + 1):
        normalized = {_normalize_header(sheet.cell(row_idx, col).value): col for col in range(1, sheet.max_column + 1)}
        field_to_col: dict[str, int] = {}
        for field, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                col = _find_header_col(normalized, alias)
                if col:
                    field_to_col[field] = col
                    break
        if "product_name_raw" in field_to_col and _has_any_quantity_field(field_to_col):
            return row_idx, field_to_col
    return None, {}


def _find_header_col(normalized_headers: dict[str, int], alias: str) -> int | None:
    normalized_alias = _normalize_header(alias)
    return normalized_headers.get(normalized_alias)


def _has_any_quantity_field(field_to_col: dict[str, int]) -> bool:
    return any(
        field in field_to_col
        for field in [
            "ordered_quantity",
            "arrival_quantity",
            "received_quantity",
            "net_quantity",
            "shipped_quantity",
            "supplied_quantity",
        ]
    )


def _choose_arrival_quantity(values: dict[str, str], ordered_quantity: float | None) -> float | None:
    candidates = [
        _to_float(values.get("received_quantity")),
        _to_float(values.get("net_quantity")),
        _to_float(values.get("shipped_quantity")),
        _to_float(values.get("supplied_quantity")),
        _to_float(values.get("arrival_quantity")),
        ordered_quantity,
    ]
    for candidate in candidates:
        if candidate is not None:
            return candidate
    return None


def _resolve_source_format(
    source_format: str,
    field_to_col: dict[str, int],
    workbook_path: Path,
    metadata: dict[str, object],
) -> str:
    if source_format != "auto":
        return source_format
    source_text = f"{workbook_path.name} {metadata.get('supplier_name', '')}"
    if "水果" in source_text:
        return "fruit_standard"
    if "蔬菜" in source_text:
        return "vegetable_supplier"
    if "net_quantity" in field_to_col or "received_quantity" in field_to_col:
        return "vegetable_supplier"
    return "vegetable_supplier"


def _normalize_header(value: object) -> str:
    text = _cell_value(value)
    text = re.sub(r"\s+", "", text)
    text = text.replace("（", "(").replace("）", ")")
    return text


def _cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _to_float(value: object) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text in {"", "-", "/"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _coerce_date(value: object) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m.%d", "%m-%d", "%m/%d"):
        try:
            parsed = datetime.strptime(text, fmt)
        except ValueError:
            continue
        year = parsed.year if "%Y" in fmt else date.today().year
        return date(year, parsed.month, parsed.day)
    return None


def _date_from_filename(filename: str) -> date | None:
    match = re.search(r"(?P<month>\d{1,2})[.\-/月](?P<day>\d{1,2})", filename)
    if not match:
        return None
    return date(date.today().year, int(match.group("month")), int(match.group("day")))


def _sheet_metadata(workbook_path: Path, sheet) -> dict[str, object]:
    text_parts: list[str] = []
    for row in range(1, min(sheet.max_row, 8) + 1):
        for col in range(1, min(sheet.max_column, 8) + 1):
            value = _cell_value(sheet.cell(row, col).value)
            if value:
                text_parts.append(value)
    text = "\n".join(text_parts)
    metadata: dict[str, object] = {}

    supplier_code = re.search(r"供应商编号[；;:：]?\s*(\d+)", text)
    if supplier_code:
        metadata["supplier_code"] = supplier_code.group(1)

    supplier_name = re.search(r"^(.+?)(?:供应商编号|$)", text)
    if supplier_name:
        name = supplier_name.group(1).strip(" ；;:：")
        if name:
            metadata["supplier_name"] = name

    store_name = re.search(r"门店名[：:]\s*(.+?)(?:\n|发货日期|采购专员|$)", text)
    if store_name:
        metadata["store_name"] = store_name.group(1).replace("丰年-", "").strip()

    ship_date = re.search(r"发货日期[：:]\s*(\d{1,2})[.\-/月](\d{1,2})", text)
    if ship_date:
        metadata["arrival_date"] = date(date.today().year, int(ship_date.group(1)), int(ship_date.group(2)))

    if "arrival_date" not in metadata:
        filename_date = _date_from_filename(workbook_path.name)
        if filename_date:
            metadata["arrival_date"] = filename_date
    return metadata


def _store_name_from_filename_or_sheet(workbook_path: Path, sheet_title: str) -> str:
    if sheet_title and sheet_title not in {"Sheet", "Sheet1", "明细", "订单"}:
        return sheet_title.strip()
    return workbook_path.stem.strip()

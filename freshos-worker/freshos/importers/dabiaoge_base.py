from __future__ import annotations

from dataclasses import asdict, dataclass
from pathlib import Path
import csv
from typing import Iterable

from openpyxl import load_workbook


DABIAOGE_BASE_COLUMNS = [
    "店铺编号", "店铺名称", "店铺状态", "大分类编码", "大分类名称", "中分类编码", "中分类名称", "小分类编码", "小分类名称",
    "品类经理", "品类主任", "商品编码", "商品名称", "商品条码", "采购组编号", "采购组名称", "新品标识", "可订标识",
    "生命周期标识", "自营商品标识", "建档日期", "配送价", "售价", "采购进价", "大仓库存成本价", "商品品牌", "规格",
    "销售单位", "箱装数", "订货批量", "全店销量ABC级别", "全店销售额ABC级别", "全店毛利额ABC级别", "全店综合ABC级别",
    "保质期限(天)", "产地", "配送类型", "商品类型", "生命周期", "商品属性", "最初销售日", "最后销售日", "最初进货日",
    "最后进货日", "大库库存数量(昨日)", "大库库存金额(昨日)", "全店库存数量(昨日)", "全店库存金额(昨日)", "主档预估毛利率",
    "店铺订货标识", "店铺收货标识", "店铺退货标识", "店铺销售标识", "店铺进价", "店铺售价", "店铺促销价", "促销标识",
    "门店销量ABC级别", "门店销售额ABC级别", "门店毛利额ABC级别", "门店综合ABC级别", "近期日均销量", "门店库存数量(昨日)",
    "门店库存金额(昨日)", "门店订货上限", "门店订货下限", "店最初进货日", "店最初销售日", "店最后进货日", "店最后销售日",
    "主档预估毛利率（店铺）",
]

DABIAOGE_BASE_ALIASES = {
    "店铺编号": ["店铺编号", "门店编号", "店铺编码", "门店编码"],
    "店铺名称": ["店铺名称", "门店名称"],
    "店铺状态": ["店铺状态", "门店状态"],
    "大分类编码": ["大分类编码", "大类编码", "一级分类编码"],
    "大分类名称": ["大分类名称", "大类名称", "一级分类名称"],
    "中分类编码": ["中分类编码", "中类编码", "二级分类编码"],
    "中分类名称": ["中分类名称", "中类名称", "二级分类名称"],
    "小分类编码": ["小分类编码", "小类编码", "三级分类编码"],
    "小分类名称": ["小分类名称", "小类名称", "三级分类名称"],
    "商品编码": ["商品编码", "产品编码", "货号"],
    "商品名称": ["商品名称", "产品名称", "品名"],
    "商品条码": ["商品条码", "条码", "国际条码"],
    "销售单位": ["销售单位", "单位", "计量单位"],
    "箱装数": ["箱装数", "包装数", "装箱数"],
    "订货批量": ["订货批量", "订货倍数", "最小订货倍数"],
    "保质期限(天)": ["保质期限(天)", "保质期(天)", "保质期限", "保质期"],
    "商品属性": ["商品属性", "商品特性"],
    "店铺订货标识": ["店铺订货标识", "门店订货标识", "可订标识"],
    "店铺销售标识": ["店铺销售标识", "门店销售标识", "可销标识"],
    "近期日均销量": ["近期日均销量", "近日日均销量", "日均销量"],
    "门店库存数量(昨日)": ["门店库存数量(昨日)", "昨日门店库存数量", "门店库存数量"],
}


@dataclass(frozen=True)
class StoreBaseRow:
    store_code: str
    store_name: str
    store_status: str

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class ProductBaseRow:
    product_code: str
    product_name: str
    barcode: str
    cat_id_01: str
    cat_name_01: str
    cat_id_02: str
    cat_name_02: str
    sale_unit: str
    fresh_attribute: str
    shelf_life_days: float | None

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class StoreProductBaseRow:
    store_code: str
    product_code: str
    store_order_status: str
    store_sale_status: str
    is_orderable: bool
    is_sellable: bool
    package_size: float | None
    order_batch_qty: float | None
    safety_stock_days: float
    recent_daily_sales: float | None
    store_stock_qty_yesterday: float | None

    def to_dict(self) -> dict[str, object]:
        return asdict(self)


@dataclass(frozen=True)
class DabiaogeBaseImportResult:
    stores: list[StoreBaseRow]
    products: list[ProductBaseRow]
    store_products: list[StoreProductBaseRow]


def parse_dabiaoge_base_csv(
    path: str | Path,
    *,
    target_cat_ids: set[str] | None = None,
) -> DabiaogeBaseImportResult:
    return parse_dabiaoge_base_file(path, target_cat_ids=target_cat_ids)


def parse_dabiaoge_base_file(
    path: str | Path,
    *,
    target_cat_ids: set[str] | None = None,
) -> DabiaogeBaseImportResult:
    target_cat_ids = target_cat_ids or {"40", "42"}
    raw_rows = _read_rows(Path(path))

    stores: dict[str, StoreBaseRow] = {}
    products: dict[str, ProductBaseRow] = {}
    store_products: dict[tuple[str, str], StoreProductBaseRow] = {}

    for row in raw_rows:
        cat_id_01 = _clean_code(row.get("大分类编码"))
        if cat_id_01 not in target_cat_ids:
            continue

        store_code = _clean_code(row.get("店铺编号"))
        product_code = _clean_code(row.get("商品编码"))
        store_name = _clean_text(row.get("店铺名称"))
        product_name = _clean_text(row.get("商品名称"))
        if not store_code or not product_code or not store_name or not product_name:
            continue

        stores.setdefault(
            store_code,
            StoreBaseRow(
                store_code=store_code,
                store_name=store_name,
                store_status=_clean_text(row.get("店铺状态")),
            ),
        )
        products.setdefault(
            product_code,
            ProductBaseRow(
                product_code=product_code,
                product_name=product_name,
                barcode=_clean_text(row.get("商品条码")),
                cat_id_01=cat_id_01,
                cat_name_01=_clean_text(row.get("大分类名称")),
                cat_id_02=_clean_code(row.get("中分类编码")),
                cat_name_02=_clean_text(row.get("中分类名称")),
                sale_unit=_clean_text(row.get("销售单位")) or "kg",
                fresh_attribute=_clean_text(row.get("商品属性")),
                shelf_life_days=_to_float(row.get("保质期限(天)")),
            ),
        )

        key = (store_code, product_code)
        store_products[key] = StoreProductBaseRow(
            store_code=store_code,
            product_code=product_code,
            store_order_status=_clean_text(row.get("店铺订货标识")),
            store_sale_status=_clean_text(row.get("店铺销售标识")),
            is_orderable=_is_enabled_status(row.get("店铺订货标识")),
            is_sellable=_is_enabled_status(row.get("店铺销售标识")),
            package_size=_to_float(row.get("箱装数")),
            order_batch_qty=_to_float(row.get("订货批量")),
            safety_stock_days=1.0,
            recent_daily_sales=_to_float(row.get("近期日均销量")),
            store_stock_qty_yesterday=_to_float(row.get("门店库存数量(昨日)")),
        )

    return DabiaogeBaseImportResult(
        stores=list(stores.values()),
        products=list(products.values()),
        store_products=list(store_products.values()),
    )


def _read_rows(path: Path) -> list[dict[str, str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        return _read_xlsx_rows(path)
    return _read_csv_rows(path)


def _read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.readline()
        fh.seek(0)
        reader: csv.DictReader[str] | csv.reader
        if _looks_like_header(sample.split(",")):
            reader = csv.DictReader(fh)
            return [_canonicalize_row(dict(row)) for row in reader]
        reader = csv.reader(fh)
        return [_row_from_values(values) for values in reader]


def _read_xlsx_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    try:
        for sheet in workbook.worksheets:
            values_iter = sheet.iter_rows(values_only=True)
            first_values = next(values_iter, None)
            if not first_values:
                continue
            if _looks_like_header(first_values):
                headers = [_clean_text(value) for value in first_values]
                rows.extend(_canonicalize_row(dict(zip(headers, values))) for values in values_iter)
            else:
                rows.append(_row_from_values(first_values))
                rows.extend(_row_from_values(values) for values in values_iter)
    finally:
        workbook.close()
    return rows


def _row_from_values(values: Iterable[object]) -> dict[str, str]:
    text_values = [_clean_text(value) for value in values]
    padded = text_values + [""] * (len(DABIAOGE_BASE_COLUMNS) - len(text_values))
    return dict(zip(DABIAOGE_BASE_COLUMNS, padded))


def _canonicalize_row(raw: dict[str, object]) -> dict[str, str]:
    normalized = {_normalize_header(key): key for key in raw}
    row = {column: "" for column in DABIAOGE_BASE_COLUMNS}
    for column in DABIAOGE_BASE_COLUMNS:
        source = normalized.get(_normalize_header(column))
        if not source:
            for alias in DABIAOGE_BASE_ALIASES.get(column, []):
                source = normalized.get(_normalize_header(alias))
                if source:
                    break
        if source:
            row[column] = _clean_text(raw.get(source))
    return row


def _looks_like_header(values: Iterable[object]) -> bool:
    normalized_values = {_normalize_header(value) for value in values}
    header_hits = 0
    for aliases in DABIAOGE_BASE_ALIASES.values():
        if any(_normalize_header(alias) in normalized_values for alias in aliases):
            header_hits += 1
    return header_hits >= 2


def _normalize_header(value: object) -> str:
    return _clean_text(value).replace(" ", "").replace("\n", "").replace("\t", "")


def _clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return "" if text in {"", "-"} else text


def _clean_code(value: object) -> str:
    text = _clean_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text


def _to_float(value: object) -> float | None:
    text = _clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _is_enabled_status(value: object) -> bool:
    text = _clean_text(value)
    if not text:
        return True
    if "不可" in text or "不允许" in text:
        return False
    return "可" in text or "正常" in text
